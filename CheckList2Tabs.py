from PyQt5.QtWidgets import *
from PyQt5 import QtCore , QtWidgets
import requests
import sys
import os
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import TestSource
import time
import win32api
import re
import xlrd
import getpass
from datetime import  datetime

files_path = []
appName = "Check Tool"


class LineEdit(QLineEdit):

    def __init__(self, title, parent):
        super().__init__(title, parent)

        self.setAcceptDrops(True)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.accept()
        else:
            event.ignore()

    def dropEvent(self, event):
        files = [str(u.toLocalFile()) for u in event.mimeData().urls()]
        for f in files:
            temp = self.text()
            temp = temp + str(f) + "\n"
            files_path.append(str(f))
            self.setText(temp)
            # self.setText(("\n"))


class Application(QWidget):

    def __init__(self):
        super().__init__()

        self.tabs = QTabWidget()
        self.tab1 = QWidget()
        self.tab2 = QWidget()
        self.tabs.addTab(self.tab1, "Checklist")
        self.tabs.addTab(self.tab2, "Data files")
        self.initUI(self.tab1)
        # self.initUIOptions(self.tab2)
        self.layout = QVBoxLayout()
        self.layout.addWidget(self.tabs)
        self.setLayout(self.layout)
        self.setWindowTitle(appName)
        self.list_document = []
        self.correspondences = {}
        self.single_check_list = []
        self.multiple_check_list = []
        self.multiple_results = []

        self.tool_version = "Check Tool 1.0"
        self.username = os.environ['USERNAME']
        self.fileFolder = "C:/Users/" + self.username + "/AppData/Local/Temp/CheckTool/"

        self.dict_function = {
            "CheckEqualValues" : 3,
            "CheckDocumentTitle" : 2,
            "CheckDocInfoParameter" : 3,
            "CheckMultipleValues" : 3,
            "CheckHyperlink" : 2,
            "CheckDocInfoOrder" : 2,
            "CountNumberOfPoints" : 3,
            "CheckIO" : 3
        }

    def initUI(self, tab):

        tab.lblUser = QLabel("USER:", tab)
        tab.lblUser.move(175, 35)
        tab.TextBoxUser = QtWidgets.QLineEdit(tab)
        tab.TextBoxUser.resize(180, 25)
        tab.TextBoxUser.move(250, 30)
        tab.TextBoxUser.setText(getpass.getuser())
        tab.TextBoxUser.setText("E518720")

        tab.lblPass = QLabel("PASSWORD:", tab)
        tab.lblPass.move(175, 70)
        tab.TextBoxPass = QtWidgets.QLineEdit(tab)
        tab.TextBoxPass.resize(180, 25)
        tab.TextBoxPass.move(250, 65)
        tab.TextBoxPass.setEchoMode((QLineEdit.Password))
        tab.TextBoxPass.setText("Cst12323")
        # tab.TextBoxPass.setText("")

        # set controls for Checklist
        tab.label1 = QLabel("Checklist", tab)
        tab.label1.move(20, 115)
        tab.edit1 = QLineEdit('', tab)
        tab.edit1.setDragEnabled(False)
        tab.edit1.move(100, 110)
        tab.edit1.resize(430, 25)
        tab.button1 = QPushButton("Browse", tab)
        tab.button1.move(540, 110)
        tab.button1.resize(50, 25)
        tab.button1.clicked.connect(self.openFileNameDialog1)

        tab.button = QPushButton("Import Checklist", tab)
        tab.button.move(250, 400)
        tab.button.clicked.connect(self.buttonGenerateClicked)

        message = ""
        tab.textbox = QTextEdit(tab)
        tab.textbox.setText(message)
        tab.textbox.move(30, 150)
        tab.textbox.resize(550, 100)
        tab.textbox.setReadOnly(True)

        self.show()

    def openFileNameDialog1(self):
        fileName1, _filter = QFileDialog.getOpenFileName(self, 'Open File', QtCore.QDir.rootPath(), '*.*')
        self.tab1.edit1.setText(fileName1)

    def buttonGenerateClicked(self):

        self.tab1.button.setEnabled(False)

        self.DocPath = self.tab1.edit1.text()
        self.Workbook = xlrd.open_workbook(self.DocPath)

        #Single_Checking - parse the functions and the parameters from the sheet
        SheetSingleChecking = self.Workbook.sheet_by_name('Single_Checking')
        for index in range(1, SheetSingleChecking.nrows):
            current_list = []
            if SheetSingleChecking.cell(index, 0).value != "":
                current_list.append(SheetSingleChecking.cell(index, 0).value)
                current_list.append(SheetSingleChecking.cell(index, 1).value)
                if SheetSingleChecking.cell(index, 1).value == "No":
                    current_list.append("NT")
                else:
                    current_list.append(SheetSingleChecking.cell(index, 2).value)
                current_list.append(SheetSingleChecking.cell(index, 3).value)
                try:
                    nr_param = self.dict_function[SheetSingleChecking.cell(index,3).value]
                    if nr_param is not None:
                        for indexCol in range(4, 4 + nr_param):
                            current_list.append(SheetSingleChecking.cell(index, indexCol).value)
                except:
                    pass
                current_list.append("")
                self.single_check_list.append(current_list)

        #Multiple_Checking - parse the functions and the parameters from sheet
        SheetMultipleChecking = self.Workbook.sheet_by_name('Multiple_Checking')
        for index in range(2, SheetMultipleChecking.nrows):
            current_list = []
            if SheetMultipleChecking.cell(index,0).value != "":
                current_list.append(SheetMultipleChecking.cell(index, 0).value)
                current_list.append(SheetMultipleChecking.cell(index, 1).value)
                if SheetMultipleChecking.cell(index, 1).value == "No":
                    current_list.append("NT")
                else:
                    current_list.append(SheetMultipleChecking.cell(index, 2).value)
                current_list.append(SheetMultipleChecking.cell(index, 3).value)
                try:
                    nr_param = self.dict_function[SheetMultipleChecking.cell(index, 3).value]
                    if nr_param is not None:
                        for indexCol in range(4, 4 + nr_param):
                            current_list.append(SheetMultipleChecking.cell(index, indexCol).value)
                except:
                    pass
                self.multiple_check_list.append(current_list)
            elif SheetMultipleChecking.cell(index, 0).value == "" and SheetMultipleChecking.cell(index + 1, 0).value == "":
                break


        #Config -- parse the documents from the Config sheet
        Sheet = self.Workbook.sheet_by_name('Config')
        docColIndex = -1
        docRowIndex = -1

        for index1 in range(0, Sheet.nrows):
            for index2 in range(0, Sheet.ncols):
                if Sheet.cell(index1, index2).value == "Documents":
                    docRowIndex = index1
                    docColIndex = index2
                    break
            if docRowIndex != -1:
                break

        for index in range(docRowIndex + 1, Sheet.nrows):
            if Sheet.cell(index, docColIndex).value != "":
                self.list_document.append(Sheet.cell(index, docColIndex).value)


        y_label = 35
        y_edit = 30
        for document_name in self.list_document:
            exec('self.tab2.label' + document_name + ' = QLabel(' + "\'" + document_name + "\'" + ', self.tab2)')
            exec('self.tab2.label' + document_name + '.move(20, ' + str(y_label) + ')')
            exec('self.tab2.edit' + document_name + ' = LineEdit(' + "\'" + "\'" + ', self.tab2)')
            exec('self.tab2.edit' + document_name + '.setDragEnabled(False)')
            exec('self.tab2.edit' + document_name + '.move(120, ' + str(y_edit) + ')')
            exec('self.tab2.edit' + document_name + '.resize(430, 25)')
            y_label += 30
            y_edit += 30

        self.tab2.button2 = QPushButton("Start Check", self.tab2)
        self.tab2.button2.move(250, 400)
        self.tab2.button2.clicked.connect(self.buttonCheckClicked)

    def download_file(self, url, user, password):

        try:
            os.stat(self.fileFolder)
        except:
            os.mkdir(self.fileFolder)
        try:
            response = requests.get(url, stream=True, auth=(user, password))
        except:
            return "Error"

        status = response.status_code
        FileName = response.headers['Content-Disposition'].split('"')[1]
        FilePath = self.fileFolder + FileName

        with open(FilePath, 'wb') as f:
            for chunk in response.iter_content(chunk_size=128):
                f.write(chunk)
        return FilePath

    def buttonCheckClicked(self):

        d1 = datetime.strptime(str(datetime.now()), "%Y-%m-%d %H:%M:%S.%f")

        self.start_time = time.time()

        for filename in self.list_document:
            exec("self.correspondences['" + filename + "'] = self.tab2.edit" + filename + ".text().strip()")

        DocPath = self.tab1.edit1.text()
        Workbook = xlrd.open_workbook(DocPath)

        print(self.correspondences)
        nr_param_max = 0

        for test in self.single_check_list:
            if test[1] == "Yes":
                if test[3] in self.dict_function:
                    nr_param = self.dict_function[test[3]]
                    if nr_param == 2 and test[4] != "" and test[5] != "":
                        if nr_param > nr_param_max:
                            nr_param_max = nr_param

                        if test[3] == 'CheckDocumentTitle':
                                test[2] = TestSource.CheckDocumentTitle(self, test[4], test[5])
                        elif test[3] == 'CheckHyperlink':
                                test[2] = TestSource.CheckHyperlink(self, test[4], test[5])
                        elif test[3] == 'CheckDocInfoOrder':
                                test[2] = TestSource.CheckDocInfoOrder(self, test[4], test[5])

                    elif nr_param == 3 and test[4] != "" and test[5] != "" and test[6] != "":
                        if nr_param > nr_param_max:
                            nr_param_max = nr_param

                        if test[3] == 'CheckEqualValues':
                            test[2], test[len(test) - 1] = TestSource.CheckEqualValues(self, test[4], test[5], test[6])
                        elif test[3] == 'CountNumberOfPoints':
                            test[2] = TestSource.CountNumberOfPoints(self, test[4], test[5], test[6])
                        elif test[3] == 'CheckDocInfoParameter':
                            test[2] = TestSource.CheckDocInfoParameter(self, test[4], test[5], test[6])

                    elif nr_param == 3 and test[4] != "" and test[6] != "":
                        if nr_param > nr_param_max:
                            nr_param_max = nr_param

                        if test[3] == 'CheckDocInfoParameter':
                            test[2] = TestSource.CheckDocInfoParameter(self, test[4], test[5], test[6])

                    else:
                        test[2] = 'NA'
                else:
                    test[2] = 'NA'

        self.vsm_sheets = []
        try:
            PathTP = self.tab2.editTP.text()
            if PathTP != "":
                PathTP = PathTP.replace("\n","")
                WorkbookTP = xlrd.open_workbook(PathTP)

                for sheet in WorkbookTP.sheet_names():
                    if re.match('^VSM.+_[0-9]{4}[A-Z]?$', sheet):
                        self.vsm_sheets.append(sheet)
        except:
            pass


        for test in self.multiple_check_list:
            if test[1] == "Yes":
                if test[3] in self.dict_function:
                    nr_param = self.dict_function[test[3]]
                    if nr_param == 3 and test[4] != "" and test[5] != "" and test[6] :
                        if test[3] == "CheckMultipleValues":
                            TestSource.CheckMultipleValues(self, test[4], test[5], test[6])
                        elif test[3] == "CheckIO":
                            TestSource.CheckIO(self, test[4], test[5], test[6])
                    else:
                        test[2] = 'NA'
                else:
                    test[2] = 'NA'

        # Write the file
        # Open the file with openpyxl (xlsm, xlsx)
        if DocPath.split(".")[-1] == "xlsm":
            WorkbookToWrite = openpyxl.load_workbook(DocPath, keep_vba = True)
        else:
            WorkbookToWrite = openpyxl.load_workbook(DocPath)

        #Single_Checking - write result in sheet
        SheetSingleChecking = WorkbookToWrite['Single_Checking']
        row = 2
        for test in self.single_check_list:
            my_cell = SheetSingleChecking.cell(row, 3)
            my_cell.value = test[2]
            my_cell = SheetSingleChecking.cell(row, 4 + nr_param_max + 1)
            my_cell.value = test[len(test) - 1]
            row += 1

        #Multiple_Checking - write result in sheet
        max_column = Workbook.sheet_by_name('Multiple_Checking').ncols
        init_max_column = max_column
        for sheet in self.vsm_sheets:
            current_cell = WorkbookToWrite['Multiple_Checking'].cell(1, max_column + 1)
            current_cell.value = sheet
            current_cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90, wrap_text=True,
                                               shrink_to_fit=False, indent=0)
            max_column += 1

        SheetMultipleChecking = WorkbookToWrite['Multiple_Checking']

        row = 3
        for list in self.multiple_results:
            i = 1
            for elem in list:
                my_cell = SheetMultipleChecking.cell(row,init_max_column + i)
                my_cell.value = elem
                i += 1
            row +=1



        WorkbookToWrite.save(DocPath)
        d2 = datetime.strptime(str(datetime.now()), "%Y-%m-%d %H:%M:%S.%f")
        print(d2 - d1)
        self.final_time = time.time()
        # self.tab1.textbox.setText("Tests duration is " + time.strftime('%H:%M:%S', time.gmtime(self.final_time - self.start_time)))
        self.tab1.textbox.setText("Tests duration is " + str(d2 - d1))

        win32api.MessageBox(0, 'Tests finished successfully!', 'Information')

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = Application()
    ex.show()
    app.exec_()
