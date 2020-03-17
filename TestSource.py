import os
import shutil
import zipfile
import openpyxl
from lxml import etree, objectify
import xlrd
from datetime import datetime
import string
import CheckList2Tabs

def column_to_number(c):
    """Return number corresponding to excel-style column."""
    number=-25
    for l in c:
        if not l in string.ascii_letters:
            return False
        number+=ord(l.upper())-64+25
    return number

def CheckEqualValues(self, Reference1, Reference2, Equal):

    try:
        t1 = datetime.strptime(str(datetime.now()), "%Y-%m-%d %H:%M:%S.%f")

        if Reference1.split("<>")[0] in self.correspondences:
            DocPath1 = self.correspondences[Reference1.split("<>")[0]]
        else:
            DocPath1 = Reference1.split("<>")[0]

        if Reference2.split("<>")[0] in self.correspondences:
            DocPath2 = self.correspondences[Reference2.split("<>")[0]]
        else:
            DocPath2 = Reference2.split("<>")[0]


        DocWorkbook1 = xlrd.open_workbook(DocPath1)
        DocWorkbook2 = xlrd.open_workbook(DocPath2)

        SheetName1 = Reference1.split("<>")[1]
        SheetName2 = Reference2.split("<>")[1]

        DocCel1 = Reference1.split("<>")[2]
        DocCel2 = Reference2.split("<>")[2]

        row1 = ""
        col1 = ""
        row2 = ""
        col2 = ""
        for char in DocCel1:
            if char.isalpha():
                col1 += char
            else:
                row1 += char

        for char in DocCel2:
            if char.isalpha():
                col2 += char
            else:
                row2 += char

        DocSheet1 = DocWorkbook1.sheet_by_name(SheetName1)
        CelValue1 = DocSheet1.cell(int(row1) - 1, column_to_number(col1)-1).value

        DocSheet2 = DocWorkbook2.sheet_by_name(SheetName2)
        CelValue2 = DocSheet2.cell(int(row2) - 1, column_to_number(col2)-1).value


        if str(Equal) == "1" or str(Equal).casefold() == "true":
            if CelValue1 == CelValue2:
                t2 = datetime.strptime(str(datetime.now()), "%Y-%m-%d %H:%M:%S.%f")
                return 'OK', str(t2 - t1)
            else:
                t2 = datetime.strptime(str(datetime.now()), "%Y-%m-%d %H:%M:%S.%f")
                return 'NOK', str(t2 - t1)

        elif str(Equal) == "0" or str(Equal).casefold() == "false":
            if CelValue1 != CelValue2:
                t2 = datetime.strptime(str(datetime.now()), "%Y-%m-%d %H:%M:%S.%f")
                return 'OK', str(t2 - t1)
            else:
                t2 = datetime.strptime(str(datetime.now()), "%Y-%m-%d %H:%M:%S.%f")
                return 'NOK', str(t2 - t1)
    except:
        return 'NA', ""

def CheckDocumentTitle(self, Reference1, Reference2):
    try:
        t1 = datetime.strptime(str(datetime.now()), "%Y-%m-%d %H:%M:%S.%f")

        if Reference1 in self.correspondences:
            Value1 = self.correspondences[Reference1].split("/")[-1]

        if Reference2.split("<>")[0] in self.correspondences:
            DocPath2 = self.correspondences[Reference2.split("<>")[0]]
        else:
            DocPath2 = Reference2.split("<>")[0]

        DocWorkbook2 = xlrd.open_workbook(DocPath2)
        DocSheet2 = DocWorkbook2.sheet_by_name(Reference2.split("<>")[1])
        col_value = ""
        row_value = ""
        for char in Reference2.split("<>")[2]:
            if char.isalpha():
                col_value += char
            else:
                row_value += char

        Value2 = DocSheet2.cell(int(row_value) - 1, column_to_number(col_value) - 1).value

        if Value1.split(".")[0] == Value2.split(".")[0]:
            t2 = datetime.strptime(str(datetime.now()), "%Y-%m-%d %H:%M:%S.%f")
            return 'OK', str(t2 - t1)
        else:
            t2 = datetime.strptime(str(datetime.now()), "%Y-%m-%d %H:%M:%S.%f")
            return 'NOK', str(t2 - t1)
    except:
        return 'NA', ""

def CheckDocInfoParameter(self, DocInfoReference, NumberFile, Parameter):
    try:
        t1 = datetime.strptime(str(datetime.now()), "%Y-%m-%d %H:%M:%S.%f")

        if NumberFile == "":
            pass
        else:
            NumberFile = str(int(NumberFile))
        if NumberFile == "":
            DocLinkIntranet = 'http://docinfogroupe.inetpsa.com/ead/doc/ref.' + DocInfoReference + '/v.vc/pj'
            DocLinkInternet = 'https://docinfogroupe.psa-peugeot-citroen.com/ead/doc/ref.' + DocInfoReference + '/v.vc/pj'
        else:
            DocLinkIntranet = 'http://docinfogroupe.inetpsa.com/ead/doc/ref.' + DocInfoReference + '/v.vc/nPj.' + NumberFile
            DocLinkInternet = 'https://docinfogroupe.psa-peugeot-citroen.com/ead/doc/ref.' + DocInfoReference + '/v.vc/nPj.' + NumberFile


        User = self.tab1.TextBoxUser.text()
        Password = self.tab1.TextBoxPass.text()
        FilePath = self.download_file(DocLinkIntranet, User, Password)
        if FilePath == "Error":
            FilePath = self.download_file(DocLinkInternet, User, Password)

        extensions = ["xlsx", "xlsm"]
        if FilePath.split(".")[-1] in extensions:
            ext = FilePath.split(".")[0]
            with zipfile.ZipFile(FilePath, 'r') as zip_ref:
                zip_ref.extractall(ext)

            try:
                if os.path.isfile(ext + "\docProps\custom.xml"):
                    path = ext + "\docProps\custom.xml"
                    parser = etree.XMLParser(remove_comments=True)
                    tree = objectify.parse(path, parser=parser)
                    root = tree.getroot()
                    returned_parameter = root.find(
                        ".//{http://schemas.openxmlformats.org/officeDocument/2006/custom-properties}property[@name = " + "\'" + Parameter + "\'" + "]/{http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes}lpwstr").text
                    shutil.rmtree(ext, ignore_errors=True)
            except:
                shutil.rmtree(ext, ignore_errors=True)
        t2 = datetime.strptime(str(datetime.now()), "%Y-%m-%d %H:%M:%S.%f")
        return returned_parameter, str(t2 -t1)
    except:
        return 'NA', ""

def CheckHyperlink(self, Hyperlink, Reference):
    try:
        t1 = datetime.strptime(str(datetime.now()), "%Y-%m-%d %H:%M:%S.%f")

        if Hyperlink.split("<>")[0] in self.correspondences:
            DocPath1 = self.correspondences[Hyperlink.split("<>")[0]]
        else:
            DocPath1 = Hyperlink.split("<>")[0]
        DocWorkbook1 = openpyxl.load_workbook(DocPath1, data_only=True, keep_vba=False)

        if Reference.split("<>")[0] in self.correspondences:
            DocPath2 = self.correspondences[Reference.split("<>")[0]]
        else:
            DocPath2 = Reference.split("<>")[0]
        DocWorkbook2 = openpyxl.load_workbook(DocPath2, data_only=True, keep_vba=False)

        SheetName1 = Hyperlink.split("<>")[1]
        SheetName2 = Reference.split("<>")[1]

        DocCel1 = Hyperlink.split("<>")[2]
        DocCel2 = Reference.split("<>")[2]

        DocSheet1 = DocWorkbook1[SheetName1]
        CelValue1 = DocSheet1[DocCel1].hyperlink.target

        DocSheet2 = DocWorkbook2[SheetName2]
        CelValue2 = DocSheet2[DocCel2].value

        if CelValue1 == CelValue2:
            t2 = datetime.strptime(str(datetime.now()), "%Y-%m-%d %H:%M:%S.%f")
            return 'OK', str(t2 - t1)
        else:
            t2 = datetime.strptime(str(datetime.now()), "%Y-%m-%d %H:%M:%S.%f")
            return 'NOK', str(t2 - t1)
    except:
        return 'NA', ""

def CheckDocInfoOrder(self, DocInfoReference, Reference):
    try:
        t1 = datetime.strptime(str(datetime.now()), "%Y-%m-%d %H:%M:%S.%f")

        DocLinkIntranet = 'http://docinfogroupe.inetpsa.com/ead/doc/ref.' + DocInfoReference + '/v.vc/pj'
        DocLinkInternet = 'https://docinfogroupe.psa-peugeot-citroen.com/ead/doc/ref.' + DocInfoReference + '/v.vc/pj'

        User = self.tab1.TextBoxUser.text()
        Password = self.tab1.TextBoxPass.text()
        FilePath = self.download_file(DocLinkIntranet, User, Password)
        if FilePath == "Error":
            FilePath = self.download_file(DocLinkInternet, User, Password)

        Doc1Name = FilePath.split("/")[-1]

        if Reference.split("<>")[0] in self.correspondences:
            DocPath2 = self.correspondences[Reference.split("<>")[0]]
        else:
            DocPath2 = Reference.split("<>")[0]

        SheetName2 = Reference.split("<>")[1]
        DocCel2 = Reference.split("<>")[2]

        DocWorkbook2 = openpyxl.load_workbook(DocPath2, data_only=True, keep_vba=False)
        DocSheet2 = DocWorkbook2[SheetName2]
        CelValue2 = DocSheet2[DocCel2].value

        if Doc1Name == CelValue2:
            t2 = datetime.strptime(str(datetime.now()), "%Y-%m-%d %H:%M:%S.%f")
            return 'OK', str(t2 -t1)
        else:
            t2 = datetime.strptime(str(datetime.now()), "%Y-%m-%d %H:%M:%S.%f")
            return 'NOK', str(t2 - t1)
    except:
        return 'NA', ""

def CountNumberOfPoints(self, Reference, Column, Equal):
    try:
        t1 = datetime.strptime(str(datetime.now()), "%Y-%m-%d %H:%M:%S.%f")

        if Reference.split("<>")[0] in self.correspondences:
            DocPath = self.correspondences[Reference.split("<>")[0]]
        else:
            DocPath = Reference.split("<>")[0]

        DocWorkbook = xlrd.open_workbook(DocPath)
        SheetName = Reference.split("<>")[1]
        DocCel = Reference.split("<>")[2]
        DocSheet = DocWorkbook.sheet_by_name(SheetName)

        col_value = ""
        row_value = ""
        for char in DocCel:
            if char.isalpha():
                col_value += char
            else:
                row_value += char

        # CelValue = DocSheet.cell(int(row_value) - 1, column_to_number(col_value) -1 ).value

        input_values = []
        try:
            column = int(Column)
            WorkSheet = self.Workbook.sheet_by_name('Config')
            for row in range(1, WorkSheet.nrows):
                if WorkSheet.cell(row, int(Column) - 1).value != "":
                    input_values.append(WorkSheet.cell(row, int(Column) -1 ).value)
        except:
            input_values.append(Column)

        number_points = 0

        row = int(row_value) - 1
        if str(Equal) == "1" or str(Equal).casefold() == "true":
            while(DocSheet.cell(row, column_to_number(col_value) -1 ).value != ""):
                if DocSheet.cell(row, column_to_number(col_value) -1 ).value in input_values:
                    number_points += 1
                row += 1

        elif str(Equal) == "0" or str(Equal).casefold() == "false":
            while (DocSheet.cell(row, column_to_number(col_value) - 1).value != ""):
                if DocSheet.cell(row, column_to_number(col_value) - 1).value in input_values:
                    number_points += 1
                row += 1
        t2 = datetime.strptime(str(datetime.now()), "%Y-%m-%d %H:%M:%S.%f")
        return number_points, str(t2 - t1)
    except:
        return 'NA', ""

def CheckMultipleValues(self, Column, Reference, Equal):

    try:
        t1 = datetime.strptime(str(datetime.now()), "%Y-%m-%d %H:%M:%S.%f")

        WorkSheet = self.Workbook.sheet_by_name('Config')
        column = int(Column)
        list_values = []
        for row in range(1, WorkSheet.nrows):
            if WorkSheet.cell(row, column - 1).value != "":
                list_values.append(WorkSheet.cell(row, column - 1).value)

        if Reference.split("<>")[0] in self.correspondences:
            DocPath = self.correspondences[Reference.split("<>")[0]]
        else:
            DocPath = Reference.split("<>")[0]
        DocWorkbook = xlrd.open_workbook(DocPath)
        current_results = []
        for sheet in self.vsm_sheets:
            DocSheet = DocWorkbook.sheet_by_name(sheet)
            row = ""
            col = ""
            for char in Reference.split("<>")[1]:
                if char.isalpha():
                    col += char
                else:
                    row += char

            Value = DocSheet.cell(int(row) - 1,column_to_number(col) - 1).value
            if str(Equal) == "1" or str(Equal).casefold() == "true":
                if Value in list_values:
                    current_results.append('OK')
                else:
                    current_results.append('NOK')
            elif str(Equal) == "0" or str(Equal).casefold() == "false":
                if Value not in list_values:
                    current_results.append('OK')
                else:
                    current_results.append('NOK')
        t2 = datetime.strptime(str(datetime.now()), "%Y-%m-%d %H:%M:%S.%f")
        current_results.append(str(t2 - t1))
        self.multiple_results.append(current_results)
    except:
        return 'NA'

def CheckIO(self, Reference1, Reference2, Type):
    try:
        t1 = datetime.strptime(str(datetime.now()), "%Y-%m-%d %H:%M:%S.%f")


        if Reference1.split("<>")[0] in self.correspondences:
            DocPath1 = self.correspondences[Reference1.split("<>")[0]]
        else:
            DocPath1 = Reference1.split("<>")[0]

        if Reference2.split("<>")[0] in self.correspondences:
            DocPath2 = self.correspondences[Reference2.split("<>")[0]]
        else:
            DocPath2 = Reference2.split("<>")[0]

        DocWorkbook1 = xlrd.open_workbook(DocPath1)
        Column1 = Reference1.split("<>")[1]

        DocWorkbook2 = xlrd.open_workbook(DocPath2)

        dci_values = []
        for sheet in DocWorkbook2.sheet_names():
            colFlux = -1
            rowFlux = -1
            colPC = -1
            rowPC = -1
            SheetName2 = DocWorkbook2.sheet_by_name(sheet)
            for index1 in range(0, SheetName2.nrows):
                for index2 in range(0, SheetName2.ncols):
                    if str(SheetName2.cell(index1, index2).value).strip() == "Flux":
                        rowFlux = index1
                        colFlux = index2
                    if str(SheetName2.cell(index1, index2).value).strip() == "P/C":
                        rowPC = index1
                        colPC = index2
                    if rowFlux != -1 and rowPC != -1:
                        break
                if rowFlux != -1 and rowPC != -1:
                    break
            if rowFlux != -1 and rowPC != -1:
                for index in range(rowFlux + 1, SheetName2.nrows):
                    if SheetName2.cell(index, colFlux).value != "":
                        dict = {}
                        dict['flux'] = str(SheetName2.cell(index, colFlux).value)
                        dict['pc'] = str(SheetName2.cell(index, colPC).value)
                        dci_values.append(dict)

        if int(Column1) == 5:
            current_results = []
            for sheet in self.vsm_sheets:
                nr_errors = 0
                DocSheet = DocWorkbook1.sheet_by_name(sheet)
                for index in range(0, DocSheet.nrows):
                    if str(DocSheet.cell(index, int(Column1) -1).value) != "" and str(DocSheet.cell(index, int(Column1) -1 ).value)[0] == "$" and str(DocSheet.cell(index, int(Column1) - 2).value) != "FONCTION":
                        flag = False
                        for elem in dci_values:
                            if str(DocSheet.cell(index, int(Column1) - 1).value)[1:] == elem['flux'] and str(Type) == elem['pc']:
                                flag = True
                        if flag is False:
                            nr_errors += 1
                if nr_errors == 0:
                    current_results.append("OK")
                else:
                    current_results.append(nr_errors)

            t2 = datetime.strptime(str(datetime.now()), "%Y-%m-%d %H:%M:%S.%f")
            current_results.append(str(t2 - t1))
            self.multiple_results.append(current_results)

        elif int(Column1) == 11:
            current_results = []
            for sheet in self.vsm_sheets:
                nr_errors = 0
                DocSheet = DocWorkbook1.sheet_by_name(sheet)
                for index in range(0, DocSheet.nrows):
                    if str(DocSheet.cell(index, int(Column1) - 1).value) != "" and str(DocSheet.cell(index, int(Column1) - 1).value)[0] == "$":
                        flag = False
                        for elem in dci_values:
                            if str(DocSheet.cell(index, int(Column1) - 1).value)[1:] == elem['flux'] and str(Type) == elem['pc']:
                                flag = True
                        if flag is False:
                            nr_errors += 1
                if nr_errors == 0:
                    current_results.append("OK")
                else:
                    current_results.append(nr_errors)

            t2 = datetime.strptime(str(datetime.now()), "%Y-%m-%d %H:%M:%S.%f")
            current_results.append(str(t2 - t1))
            self.multiple_results.append(current_results)

    except:
        return 'NA'
